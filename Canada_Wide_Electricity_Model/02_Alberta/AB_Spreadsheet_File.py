#!/usr/bin/env python
# -*- coding: latin-1 -*-
"""
    Support for the Alberta spreadsheet input.  

    Note that this spreadsheet has detailed information
    for each power generating station.
"""

from optparse import OptionParser
from collections import OrderedDict
import operator
import sys
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from datetime import datetime, timezone

curdir = os.path.dirname(os.path.abspath(__file__))
if curdir not in sys.path:
    sys.path.append(curdir)
comdir = os.path.join(os.path.split(curdir)[0], "Common")
if comdir not in sys.path:
    sys.path.append(comdir)

from common_defs import *
from demand_file import DemandFile
import generator_file
from hourly_mw_file import HourlyMWFile

class AB_Spreadsheet_File(object):
    def __init__(self, file_paths = [], asset_paths = []):
        self.lines = []
        self.load_lines = []
        self.files = []
        self.assets = {}
        self.demand_file = DemandFile()

        self.gen_file = generator_file.generator_file()
        for path in asset_paths:
            self.check_file(path)
            self.read_xls_file(path)
            self.parse_asset_list(path)

        self.fuel_gen_db = {}
        for fuel in self.gen_file.gen_db.keys():
            self.fuel_gen_db[fuel] = HourlyMWFile()

        for path in file_paths:
            try:
                self.check_file(path)
                logging.info("Reading '%s'." % path)
                self.read_xls_file(path)
                self.load_lines.append(self.lines)
                logging.info("Parsing '%s'." % path)
                self.parse_load_file(path)
                self.files.append(path)
                logging.info("Parsed file %s" % path)
            except ValueError as e:
                if do_all_files:
                    continue
                print(e)
                sys.exit(-1)

    def check_file(self, file_path):
        if not os.path.isfile(file_path):
            raise ValueError("File '%s' does not exist!" % file_path)
        filename, file_extension = os.path.splitext(file_path)
        if EXCEL_FILE_EXTENSION != file_extension:
            raise ValueError("File '%s' wrong type, want '%s' but got %s!" %
                    (file_path, EXCEL_FILE_EXTENSION, file_extension))

    def read_excel(self, file_path):
        self.wb = load_workbook(filename = file_path)
        for sheet in self.wb:
            if sheet.title != self.wb.active.title:
                continue
            idx = 0
            for row in sheet.iter_rows(min_row = 1):
               vals = [c.value  for c in row]
               for i, val in enumerate(vals):
                   if val is None:
                       vals[i] = ''
                   try:
                       vals[i] = str(vals[i])
                   except UnicodeEncodeError:
                       vals[i].encode("ascii","ignore")
               line = SEPARATOR.join(vals)
               line_crlf = "%s%s%s\n" % (START_END, line, START_END)
               self.lines.append(line_crlf)
               idx += 1
        self.wb.close()

    def _parse_date(self, line_num, token):
        month_abbrev = ["SKIP", "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
                                "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
        try:
            day = int(token[0:2])
            month = month_abbrev.index(token[2:5])
            year = int(token[5:9])
            hour = int(token[10:12])
        except ValueError as e:
            logging.critical("Could not parse line %d date '%s'.  Halting." % (line_num, token))
            exit()

        return year, month, day, hour

    def parse_load_file(self, path):
        if self.lines is []:
            logging.error("%s is empty!" % path)
            return
        logging.debug("\nFirst 5 lines\n%s\n" % self.lines[0:5])
        header = self.lines[0].strip()
        toks = [tok.strip() for tok in header[1:-1].split(SEPARATOR)]
        logging.debug("\nHeader tokens %s\n" % toks)

        try:
            GMT_index = toks.index("Date_Begin_GMT")
            loc_index = toks.index("Date_Begin_Local")
            load_index = toks.index("ACTUAL_AIL")
            fuel_list = []
            for tok in toks:
                if tok in self.assets:
                    fuel_list.append(self.assets[tok])
                else:
                    fuel_list.append('')

        except ValueError as e:
            logging.critical("Could not find at least one of GMT, Local time, or AIL index.  Halting.")
            exit()

        for line_num, line in enumerate(self.lines[1:]):
            line = line.strip()
            toks = [tok.strip() for tok in line[1:-1].split(SEPARATOR)]
            try:
                GMT_year, GMT_month, GMT_day, GMT_hour = self._parse_date(line_num, toks[GMT_index])
                loc_year, loc_month, loc_day, loc_hour = self._parse_date(line_num, toks[loc_index])
                the_load = float(toks[load_index])
                new_tuple = [path, line_num,
                             GMT_year, GMT_month, GMT_day, GMT_hour,
                             loc_year, loc_month, loc_day, loc_hour, the_load]
                self.demand_file.add_mw_hour(path, line_num, new_tuple)
                UMT = datetime(GMT_year, GMT_month, GMT_day, hour=GMT_hour)
                for idx, tok in enumerate(toks):
                    if tok == '' or fuel_list[idx] == '':
                        continue
                    ull = new_tuple[2:-1]
                    ull.append(tok)
                    self.fuel_gen_db[fuel_list[idx]].add_mw_hour(path, line_num, ull)

            except ValueError as e:
                logging.warning("Error processing %s Line %d:%s" %
                        (path, line_num, line))
                logging.warning(e)
                continue

    def write_fuel_gen_file(self, fuel, path):
        try:
            self.fuel_gen_db[fuel].write_hourly_mw_file(path)
        except:
            pass

    def read_xls_file(self, path):
        self.lines = []
        self.read_excel(path)

    def parse_asset_list(self, path):
        title = ['ASSET_SHORT_NAME', 'ASSET_NAME', 'FUEL_TYPE', 'SUB_FUEL_TYPE', 'MAXIMUM_CAPABILITY']
        line = self.lines[0].strip()
        toks = [tok.strip() for tok in line[1:-1].split(SEPARATOR)]
        if False in [x == y for x, y in zip(title, toks)]:
            logging.critical("Asset list file '%s' had unexpected title:\n'%s'\nWill not create actual generation files."
                             % (path, ' '.join(self.lines[0])))
            self.pv_solar_path = ''
            self.wind_path = ''
            return

        for li_no, line in enumerate(self.lines[1:]):
            line = line.strip()
            toks = [tok.strip() for tok in line[1:-1].split(SEPARATOR)]
            if "Solar" in toks[1]:
                fuel = "SOLAR_PV"
            elif toks[3] != '':
                fuel = find_fuel(toks[3])
            else:
                fuel = find_fuel(toks[2])
            if fuel == '':
                logging.warning("File %s line %d Could not determine fuel: '%s'"
                              % (path, li_no+1, line))
                continue
            self.assets[toks[0]] = fuel
            self.gen_file.add_generator([fuel, toks[4], '', ''])

def create_parser():
    parser = OptionParser(description="Support for Alberta Hourly Load Detailed Excel file.")
    parser.add_option('-x', '--excel',
            dest = 'excel_file_paths',
            action = 'append', type = 'string', default = [],
            help = 'File path to Alberta format Excel file.',
            metavar = 'FILE')
    parser.add_option('-a', '--assets',
            dest = 'asset_file_paths',
            action = 'append', type = 'string', default = [],
            help = 'File path to Excel list of generating assets.',
            metavar = 'FILE')
    parser.add_option('-g', '--gen_file',
            dest = 'gen_file_from_assets',
            action = 'store', type = 'string', default = '',
            help = 'File path for gen_file derived from assets file .',
            metavar = 'FILE')
    parser.add_option('-p', '--pv_solar',
            dest = 'actual_pv_solar_gen',
            action = 'store', type = 'string', default = '',
            help = 'File path for actual_pf_solar_gen file.',
            metavar = 'FILE')
    parser.add_option('-w', '--wind',
            dest = 'actual_wind_gen',
            action = 'store', type = 'string', default = '',
            help = 'File path for actual_wind_gen file.',
            metavar = 'FILE')
    return parser

def main(argv = None):
    logging.basicConfig(level=logging.INFO)
    parser = create_parser()
    if argv is None:
        argv = sys.argv[1:]

    (options, argv) = parser.parse_args(argv)
    if len(argv) != 0:
        print('Must enter at least one file path!')
        print
        parser.print_help()
        return -1

    ssheet = AB_Spreadsheet_File(options.excel_file_paths,
                                 options.asset_file_paths)
    ssheet.demand_file.write_hourly_mw_file()
    if options.asset_file_paths!= []:
        if options.gen_file_from_assets != '':
            ssheet.gen_file.write_generator_file(options.gen_file_from_assets)
        if options.actual_pv_solar_gen != '':
            ssheet.write_fuel_gen_file("SOLAR_PV", options.actual_pv_solar_gen)
        if options.actual_wind_gen != '':
            ssheet.write_fuel_gen_file("WIND", options.actual_wind_gen)

if __name__ == '__main__':
    sys.exit(main())
