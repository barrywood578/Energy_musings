#!/usr/bin/env python
# -*- coding: latin-1 -*-
"""
    Support for British Columbia XLSX spreadsheet.

    Computes UTC time for each local date/time,
    and outputs a load file in standard format.
"""

from optparse import OptionParser
from collections import OrderedDict
import operator
import sys
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from datetime import datetime, timezone
import pytz

sys.path.append('../Common')
from common_defs import *
from demand_file import *

class BCSpreadsheetFiles(object):
    def __init__(self, do_all_files = False, file_paths = []):
        self.lines = []
        self.files = []
        self.demand_file = demand_file()
        self.dst = False

        if do_all_files:
            file_paths = []
            file_list = [f for f in os.listdir('.') if os.path.isfile(f)]
            for each_file in file_list:
                filename, file_extension = os.path.splitext(each_file)
                if file_extension == EXCEL_FILE_EXTENSION:
                    file_paths.append(each_file)

        for path in file_paths:
            try:
                self.check_file(path)
                self.read_and_parse_file(path)
                self.files.append(path)
                logging.warning("Parsed file %s" % path)
            except ValueError as e:
                if do_all_files:
                    continue
                print(e)
                sys.exit(-1)

    def check_file(self, file_path):
        if not os.path.isfile(file_path):
            raise ValueError("File '%s' does not exist!" % file_path)
        filename, file_extension = os.path.splitext(file_path)
        if EXCEL_FILE_EXTENSION != file_extension:
            raise ValueError("File '%s' wrong type, want '%s' but got %s!" %
                    (file_path, EXCEL_FILE_EXTENSION, file_extension))

    def read_excel(self, file_path):
        self.wb = load_workbook(filename = file_path)
        for sheet in self.wb:
            if sheet.title != self.wb.active.title:
                continue
            idx = 0
            for row in sheet.iter_rows(min_row = 1):
               vals = [c.value  for c in row]
               for i, val in enumerate(vals):
                   if val is None:
                       vals[i] = ''
                   try:
                       vals[i] = str(vals[i])
                   except UnicodeEncodeError:
                       vals[i].encode("ascii","ignore")
               line = SEPARATOR.join(vals[0:3])
               line_crlf = "%s%s%s\n" % (START_END, line, START_END)
               self.lines.append(line_crlf)
               logging.info("Line from Excel %d: %s" % (idx, line_crlf))
               idx += 1

    def _get_BC_UTC(self, year, month, day, hour_in, dst_in=False):
        local = pytz.timezone("America/Vancouver")
        naive = datetime(year, month, day, hour=hour_in)
        local_dt = local.localize(naive, is_dst=dst_in)
        utc_dt = local_dt.astimezone(pytz.utc)
        return utc_dt.year, utc_dt.month, utc_dt.day, utc_dt.hour

    def parse_lines(self, path):
        if self.lines is []:
            logging.error("%s is empty!" % path)
            return

        self.dst = False
        prev_hour = -1

        for line_num, line in enumerate(self.lines[1:]):
            line = line.strip()
            toks = [tok.strip() for tok in line[1:-1].split(SEPARATOR)]
            if len(toks) < 3 or '' in toks[0:3] or "HE" in toks:
                logging.info("%s: Skipping line %d:'%s'" %
                        (path, line_num, line))
                continue
            try:
                the_date = [tok.strip() for tok in toks[0].split(" ")]
                try:
                    year, month, day = [int(tok.strip()) for tok in the_date[0].split("-")]
                except ValueError:
                    month, day, year = [int(tok.strip()) for tok in the_date[0].split("/")]
                if toks[1][-1] == '*':
                    toks[1] = toks[1][:-1]
                the_hour = int(toks[1]) - 1
                the_load = float(toks[2])

                ## Now for some messing about due to Daylight Savings Time entry and exit
                ##
                ## A load value of 0 indicates hour skipped at the start of daylight savings time...
                ## The load is zero, so skip it.
                if (the_load == 0.0):
                    self.dst = True
                    continue
                ## If the time is discontiguous, this is another indication that an hour was skipped
                ## at the start of daylight savings time.  Check for precisely 2 hours, to avoid
                ## obvious rollover at 24.
                if ((the_hour - prev_hour) == 2):
                    self.dst = True

                ## The second instance of an hour indicates that DST has ended.
                if (the_hour == prev_hour):
                    self.dst = False

                utc_y, utc_m, utc_d, utc_h = self._get_BC_UTC(year, month, day, the_hour, self.dst)
                prev_hour = the_hour

                self.demand_file.add_demand_hour([path, line_num,
                                                  utc_y, utc_m, utc_d,    utc_h,
                                                  year , month,   day, the_hour, the_load])
            except ValueError as e:
                logging.warning("Error processing %s Line %d:%s" %
                        (path, line_num, line))
                logging.warning(e)
                continue

    def read_and_parse_file(self, path):
        self.lines = []
        self.read_excel(path)
        self.parse_lines(path)

    def print_demand_file(self):
        self.demand_file.write_demand_file()

def create_parser():
    parser = OptionParser(description="Support for British Columbia Hourly Load Excel files.")
    parser.add_option('-x', '--excel',
            dest = 'excel_file_paths',
            action = 'append', type = 'string', default = [],
            help = 'File path to British Columbia format Excel file.',
            metavar = 'FILE')
    parser.add_option('-a', '--all',
            dest = 'all_xlsx_files',
            action = 'store_true', default = False,
            help = 'Process all .xlsx files in this directory.',
            metavar = 'FLAG')
    return parser

def main(argv = None):
    logging.basicConfig(level=logging.WARN)
    parser = create_parser()
    if argv is None:
        argv = sys.argv[1:]

    (options, argv) = parser.parse_args(argv)
    if len(argv) != 0:
        print('Must enter at least one file path!')
        print
        parser.print_help()
        return -1

    ssheet = BCSpreadsheetFiles(options.all_xlsx_files,
                                options.excel_file_paths)
    ssheet.print_demand_file()

if __name__ == '__main__':
    sys.exit(main())
